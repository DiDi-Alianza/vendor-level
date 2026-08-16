# -*- coding: utf-8 -*-
"""
历史月份（2026-04 / 05 / 06）指标重取 —— BUG-003 修正后口径。

背景：`商分级结果.xlsx` 的 `Perfect order` 列名不符实（实为完单量）且仅部分商有值，
原脚本用 `or 0` 把空值静默判为 0，权重 30% 的指标大面积错误。见 docs/指标取数口径.md §2.1b。

本脚本只取**账单口径可得**的指标：
  orders   日均完美单量 = courier by day.Perfect_Orders 自然月合计 ÷ 当月自然日数
  d3r      D-3R% = (Reject+Reschedule+Reassign) ÷ Assigned_Orders
  couriers 完单骑手数（去重）
  newrider 行业全新骑手 ÷ 当期新招骑手
  blocked  反作弊红线封禁骑手（去重） ÷ 完单骑手

⛔ 不产出 slot / credit：4–6 月的 Slot 达成率与欠款周快照源文件不存在（Slot 自 wk0629 起、
   欠款自 0622 起）。缺就是缺，不估算、不外推 —— 因此这几个月**无法出完整评级**。

用法：PYTHONIOENCODING=utf-8 python scripts/extract_history.py
产出：data/_monthly_values_2026-04.json / -05 / -06（values 里 slot、credit 为 null）
"""
import openpyxl, io, json, os, calendar, unicodedata
from collections import defaultdict

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = r"C:\Users\abbychang\Desktop\商分级\01_原始数据"
BILLS = os.path.join(SRC, "周度账单&成本", "周度账单")

# 每月覆盖该月全部自然日所需的账单（周区间可跨月，脚本按 Date 逐行过滤）
MONTHS = {
    "2026-04": ["0406Bill-wk0330_adj.xlsx", "0413Bill-wk0406_adj.xlsx", "0420Bill-wk0413_adj.xlsx",
                "0427Bill-wk0420_adj.xlsx", "0504Bill-wk0427_adj.xlsx"],
    "2026-05": ["0504Bill-wk0427_adj.xlsx", "0511Bill-wk0504_adj.xlsx", "0518Bill-wk0511_adj.xlsx",
                "0525Bill-wk0518_adj.xlsx", "0601Bill-wk0525_adj.xlsx"],
    "2026-06": ["0601Bill-wk0525_adj.xlsx", "0608Bill-wk0601_adj.xlsx", "0615Bill-wk0608_adj.xlsx",
                "0622Bill-wk0615_adj.xlsx", "0629Bill-wk0622_adj.xlsx", "0706Bill-wk0629_adj.xlsx"],
}
NEWRIDER = os.path.join(SRC, "骑手拉新", "456月新骑手.xlsx")
BLOCKS = os.path.join(SRC, "人证合一", "5-8月封禁明细.xlsx")


def norm(s):
    s = str(s).strip().upper()
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def num(x):
    if x is None: return None
    try: return float(str(x).replace("%", "").replace(",", "").strip())
    except ValueError: return None


def is_agg(c):
    u = str(c).strip().upper()
    return u.startswith("#") or "TOTAL" in u or "合计" in u or u in ("VENDOR_CODE", "CÓDIGO")


def load_bills(period, files, issues):
    y, m = int(period[:4]), int(period[5:7])
    days = calendar.monthrange(y, m)[1]
    agg = defaultdict(lambda: dict(perfect=0.0, complete=0.0, assigned=0.0, r3=0.0, couriers=set()))
    seen_days, cols = set(), set()
    for f in files:
        p = os.path.join(BILLS, f)
        if not os.path.exists(p):
            issues.append({"type": "bill_missing", "file": f}); continue
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        if "courier by day" not in wb.sheetnames:
            issues.append({"type": "no_courier_sheet", "file": f}); wb.close(); continue
        ws = wb["courier by day"]; ws.reset_dimensions(); it = ws.iter_rows(values_only=True)
        h = [str(x).strip() if x else "" for x in next(it)]
        jx = {n: i for i, n in enumerate(h) if n}
        # 完美单列名带口径后缀（(10min) / (1min)），按前缀定位，不写死
        pc = [n for n in jx if n.startswith("Perfect_Orders")]
        if len(pc) != 1:
            issues.append({"type": "perfect_col_ambiguous", "file": f, "found": pc}); wb.close(); continue
        i_perf = jx[pc[0]]; cols.add(pc[0])
        for r in it:
            if not r or r[jx["Vendor_Code"]] is None: continue
            d = str(r[jx["Date"]])[:10]
            if not (d[:4] == period[:4] and d[5:7] == period[5:7]): continue
            seen_days.add(d)
            a = agg[norm(r[jx["Vendor_Code"]])]
            a["perfect"] += num(r[i_perf]) or 0
            a["complete"] += num(r[jx["Complete_Orders"]]) or 0
            a["assigned"] += num(r[jx["Assigned_Orders"]]) or 0
            a["r3"] += sum(num(r[jx[k]]) or 0 for k in ("D_Duty_Reject", "D_Duty_Reschedule", "D_Duty_Reassign"))
            cid = r[jx["Courier_ID"]]
            if cid is not None and (num(r[jx["Complete_Orders"]]) or 0) > 0:
                a["couriers"].add(str(cid).strip())
        wb.close()
    if len(seen_days) != days:
        issues.append({"type": "days_incomplete", "expected": days, "got": len(seen_days),
                       "missing": sorted({f"{period}-{i:02d}" for i in range(1, days + 1)} - seen_days)})
    if len(cols) > 1:
        issues.append({"type": "perfect_caliber_mixed", "cols": sorted(cols),
                       "note": "当期内完美单口径发生切换，月度汇总跨口径"})
    return agg, days, seen_days, sorted(cols)


def load_newrider(period, issues):
    """456月新骑手.xlsx：单 sheet 装三个月，靠 `period` 列（如 2026-04-01~2026-04-30）区分。"""
    out = {}
    if not os.path.exists(NEWRIDER):
        issues.append({"type": "newrider_file_missing", "file": NEWRIDER}); return out
    wb = openpyxl.load_workbook(NEWRIDER, read_only=True, data_only=True)
    ws = wb.worksheets[0]; ws.reset_dimensions(); it = ws.iter_rows(values_only=True)
    h = [str(x).strip() if x else "" for x in next(it)]
    jx = {n: i for i, n in enumerate(h) if n}

    def col(*cands):
        for c in cands:
            for k in jx:
                if k.lower() == c.lower(): return jx[k]
        for c in cands:
            for k in jx:
                if c.lower() in k.lower(): return jx[k]
        return None

    i_bn, i_nr = col("brand_new_curp", "brand_new", "行业全新"), col("new_riders", "新骑手")
    i_v, i_p = col("vendor_code", "vendor"), col("period")
    if None in (i_bn, i_nr, i_v, i_p):
        issues.append({"type": "newrider_cols_missing", "period": period, "header": h})
        wb.close(); return out
    rows_seen = 0
    for r in it:
        if not r or r[i_v] is None or is_agg(r[i_v]): continue
        # period 形如 2026-04-01~2026-04-30，只取本月
        if not str(r[i_p]).startswith(period): continue
        rows_seen += 1
        out[norm(r[i_v])] = dict(brand_new=num(r[i_bn]) or 0, new_riders=num(r[i_nr]) or 0)
    wb.close()
    if rows_seen == 0:
        issues.append({"type": "newrider_period_empty", "period": period,
                       "note": "period 列没有匹配本月的行，拉新指标全为 null"})
    return out


def load_blocks(period, issues):
    """反作弊红线封禁，骑手 ID 去重，按封禁日期落月。"""
    out = defaultdict(set)
    if not os.path.exists(BLOCKS):
        issues.append({"type": "blocks_file_missing"}); return out
    wb = openpyxl.load_workbook(BLOCKS, read_only=True, data_only=True)
    ws = wb.worksheets[0]; ws.reset_dimensions(); it = ws.iter_rows(values_only=True)
    h = [str(x).strip() if x else "" for x in next(it)]
    jx = {n: i for i, n in enumerate(h) if n}

    def col(*cands):
        for c in cands:
            for k in jx:
                if c.lower() in k.lower(): return jx[k]
        return None

    i_v, i_c = col("vendor_code", "vendor"), col("courier_id", "骑手id", "rider")
    i_ch, i_bs = col("channel_source"), col("block_strategy_name")
    i_d = col("date", "时间", "封禁")
    if None in (i_v, i_c):
        issues.append({"type": "blocks_cols_missing", "header": h}); wb.close(); return out
    for r in it:
        if not r or r[i_v] is None: continue
        if i_ch is not None and "ANTI-FRAUD" not in str(r[i_ch]).upper(): continue
        if i_bs is not None and "红线" not in str(r[i_bs]): continue
        if i_d is not None:
            d = str(r[i_d])[:10]
            if not (d[:4] == period[:4] and d[5:7] == period[5:7]): continue
        out[norm(r[i_v])].add(str(r[i_c]).strip())
    wb.close()
    return out


for period, files in MONTHS.items():
    issues = []
    agg, days, seen, cols = load_bills(period, files, issues)
    NEW = load_newrider(period, issues)
    BAN = load_blocks(period, issues)
    vendors = []
    for v, a in sorted(agg.items()):
        cc = len(a["couriers"])
        n = NEW.get(v)
        vendors.append({
            "vendor_code": v, "vendor_key": v,
            "city": "MTY" if "MTY" in v else "CDMX",
            "values": {
                "orders": round(a["perfect"] / days, 2),
                "slot": None,        # ⛔ 源数据不存在，见文件头
                "credit": None,      # ⛔ 源数据不存在，见文件头
                "d3r": (round(a["r3"] / a["assigned"] * 100, 3) if a["assigned"] else None),
                "newrider": (round(n["brand_new"] / n["new_riders"] * 100, 2)
                             if n and n["new_riders"] else None),
                "blocked_rider_rate": (round(len(BAN.get(v, ())) / cc * 100, 2) if cc else None),
            },
            "raw": {
                "perfect_orders": round(a["perfect"]), "completed_orders": round(a["complete"]),
                "days_in_period": days, "assigned_orders": round(a["assigned"]),
                "d3r_events": round(a["r3"]), "completing_couriers": cc,
                "new_riders": (n["new_riders"] if n else None),
                "brand_new_curp": (n["brand_new"] if n else None),
                "blocked_riders_dedup": len(BAN.get(v, ())),
            },
        })
    out = {
        "meta": {
            "period": period, "period_type": "monthly", "days_in_period": days,
            "period_label": f"{period[:4]} 年 {int(period[5:7])} 月（自然月）",
            "vendor_count": len(vendors),
            "partial": True,
            "unavailable_indicators": ["slot", "credit"],
            "unavailable_reason": "Slot 达成率源文件自 wk0629 起、欠款周快照自 0622 起，4–6 月不存在。"
                                  "缺数不估算、不外推，因此本期无法计算综合分与等级。",
            "caliber": {
                "orders": f"周账单 courier by day.Perfect_Orders 自然月合计 ÷ {days}（列 {cols}）",
                "d3r": "(Reject+Reschedule+Reassign) ÷ Assigned_Orders",
                "newrider": "行业全新骑手(CURP) ÷ 当期新招骑手",
                "blocked_rider_rate": "反作弊红线封禁骑手(去重) ÷ 完单骑手",
            },
            "days_covered": len(seen),
            "change_request": "CR-20260814 / BUG-003",
        },
        "issues": issues,
        "vendors": vendors,
    }
    p = os.path.join(ROOT, "data", f"_monthly_values_{period}.json")
    io.open(p, "w", encoding="utf-8").write(json.dumps(out, ensure_ascii=False, indent=1))
    print(f"[{period}] {len(vendors)} 家｜覆盖 {len(seen)}/{days} 天｜完美单列 {cols}｜issues {len(issues)}")
    for x in issues[:6]:
        print("    ⚠️", json.dumps(x, ensure_ascii=False)[:150])
    print(f"    → {p}")
