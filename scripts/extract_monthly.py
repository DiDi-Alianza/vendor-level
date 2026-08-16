# -*- coding: utf-8 -*-
"""
月度取数层（自然月口径 · CR-20260814 新口径）—— 只取数，不算分
================================================================
口径（全部对齐自然月，见 docs/指标取数口径.md）：
  orders   日均完美单量 = 商分级结果.xlsx 当月 Perfect order ÷ 当月自然日数
  slot     月度达成率  = Σ当月每日达成 slot 数 ÷ Σ当月每日目标 slot 数（日表 slot ach final，阈值 70% 已验证 113/113 吻合）
  credit   逾期>7天占比 = 当月各周快照的均值（★坏账占比不再参与，CR-20260814 B1）
  d3r      D-3R%      = Σ(Reject+Reschedule+Reassign) ÷ Σ Assigned_Orders（日表按自然月过滤，公式误差 0.000pp 已验证）
  newrider 新骑手占比  = brand_new_curp ÷ new_riders（★分母由完单骑手改为新招骑手，CR-20260814 B3）
  blocked  合规账号率  = 当月 AF 红线封禁去重骑手 ÷ 当月完单骑手（日表 Courier_ID 去重）

用法：改 MONTH 配置后 `python scripts/extract_monthly.py`
"""
import openpyxl, csv, io, json, unicodedata, os, calendar
from collections import defaultdict

MONTH = dict(
    period="2026-07",
    label="2026 年 7 月（7/1–7/31 自然月）",
    year=2026, month=7,
    root=r"C:\Users\abbychang\Desktop\商分级\01_原始数据",
    bills=["0706Bill-wk0629_adj.xlsx", "0713Bill-wk0706-adj.xlsx", "0720Bill-wk0713_adj.xlsx",
           "0727Bill-wk0720_adj.xlsx", "0803Bill-wk0727 - adj.xlsx"],
    slots=["Slot_detail_0629.xlsx", "Slot_Result wk0706.xlsx", "Slot wk0713.xlsx",
           "slot_detail_0720.xlsx", "slot_detail0727.xlsx"],
    debts=[r"欠款情况\0629-0705.xlsx", r"欠款情况\0706-0712\逾期指标_0715.csv",
           r"欠款情况\0713-0719\逾期指标_0721.csv", r"欠款情况\0720-0726\逾期明细_0729.csv"],
    newrider=r"骑手拉新\7月新骑手.xlsx",
    blocks=r"人证合一\5-8月封禁明细.xlsx",
    result_sheet="JULY",
)
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DAYS = calendar.monthrange(MONTH["year"], MONTH["month"])[1]
PREFIX = f"{MONTH['year']}-{MONTH['month']:02d}"

_rules = json.load(io.open(os.path.join(ROOT, "data", "rules.json"), encoding="utf-8"))
_KEYS = [i["key"] for i in _rules["indicators"]]
def _key(*c):
    for k in c:
        if k in _KEYS: return k
    raise SystemExit(f"rules.json 无此指标 key: {c}，现有 {_KEYS}")
K_ORD, K_SLOT, K_CRED = _key("orders"), _key("slot"), _key("credit")
K_D3R, K_NEW, K_BLK = _key("d3r", "r2"), _key("newrider"), _key("blocked_rider_rate", "identity")

def norm(s):
    s = str(s).strip().upper()
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
def num(x):
    try: return float(str(x).replace("%", "").replace(",", ""))
    except: return None
def pct(x):
    v = num(x)
    return None if v is None else (v * 100 if abs(v) <= 1.5 else v)
def is_agg(c):
    u = str(c).strip().upper()
    return u.startswith("__") or u in ("TOTAL", "合计") or "TOTAL" in u
def in_month(datestr):
    return str(datestr)[:7] == PREFIX

R = MONTH["root"] + "\\"
issues = []

# ---------- 1. orders：自然月完美单 ÷ 自然日数 ----------
wb = openpyxl.load_workbook(R + "商分级结果.xlsx", read_only=True, data_only=True)
ws = wb[MONTH["result_sheet"]]
rows = ws.iter_rows(values_only=True)
h = [str(x).strip() if x else "" for x in next(rows)]
ix = {n: i for i, n in enumerate(h) if n}
V = {}
# ⚠️ BUG-003（2026-08-14）：原实现用 `Perfect order` 列并写 `or 0`，但该列源表只填了 37 家（大商），
# 另 65 家为空 → 被静默判为 0 单/天，权重 30% 的指标对 64% 的商全错。
# `Daily Completed Orders` 列 102 家全有值，且在两者都有值的 37 家上 Daily×31 与 Perfect 偏差中位 0.33%。
# 因此改用 Daily Completed Orders 作为日均单量的取数列；缺值一律 None 并入 issues，绝不补 0。
for r in rows:
    if not r or r[0] is None or is_agg(r[0]): continue
    v = norm(r[0])
    lvl = str(r[ix["July Level"]]).strip() if r[ix["July Level"]] else ""
    daily = num(r[ix["Daily Completed Orders"]])
    pm = num(r[ix["Perfect order"]])          # 仅作交叉校验，可能为空
    if daily is None:
        issues.append(f"orders 缺数（Daily Completed Orders 为空）：{r[0]}")
    elif pm is not None and pm > 0 and abs(daily * DAYS - pm) / pm > 0.05:
        issues.append(f"orders 交叉校验偏差 >5%：{r[0]} daily×{DAYS}={daily*DAYS:.0f} vs Perfect={pm:.0f}")
    V[v] = dict(code=str(r[0]).strip(), daily=daily, perfect_month=pm,
                official_level=(lvl if lvl not in ("-", "", "Sin puntos") else None),
                city="MTY" if "MTY" in v else "CDMX")
_miss = sum(1 for d in V.values() if d["daily"] is None)
print(f"[1/6] orders: {len(V)} 家（Daily Completed Orders 直取；缺数 {_miss} 家）")

# ---------- 2. slot：日表按自然月聚合 ----------
sl = defaultdict(lambda: [0, 0])   # [达成, 总数]
slot_dates = set()
for f in MONTH["slots"]:
    p = R + r"周度账单&成本\Slot达成率" + "\\" + f
    if not os.path.exists(p):
        issues.append({"type": "slot_file_missing", "file": f}); continue
    w = openpyxl.load_workbook(p, read_only=True, data_only=True)
    nm = next((x for x in w.sheetnames if "detail" in x.lower() and "rider" not in x.lower()), None)
    if not nm: w.close(); issues.append({"type": "slot_sheet_missing", "file": f}); continue
    ws = w[nm]; ws.reset_dimensions(); rows = ws.iter_rows(values_only=True)
    hh = [str(x).strip() if x else "" for x in next(rows)]
    jx = {n: i for i, n in enumerate(hh) if n}
    ivd = jx.get("Vendor Code") or jx.get("Vendor code")
    idt, itg = jx.get("Date"), jx.get("Target")
    iach = jx.get("achieve_rider_cnt")
    # ★ 判定列用 "slot ach?"（0/1 真实达成），不可用 "slot ach final"
    #   后者是含周末/比赛加权的计数（1.0/1.5/3.5/4.5），只影响全职奖金额，不进等级考核
    ifin = jx.get("slot ach?")
    if ivd is None or idt is None or itg is None: w.close(); continue
    for r in rows:
        if not r or r[ivd] is None or is_agg(r[ivd]): continue
        d = str(r[idt])[:10]
        if not in_month(d): continue
        slot_dates.add(d)
        v = norm(r[ivd]); tgt = num(r[itg])
        if tgt is None or tgt <= 0: continue
        sl[v][1] += 1
        if ifin is not None and r[ifin] is not None:
            ok = num(r[ifin])
            hit = (ok == 1) if ok is not None else (str(r[ifin]).strip().upper() in ("Y", "YES", "TRUE"))
        else:
            ach = num(r[iach]) or 0
            hit = (ach / tgt) >= 0.70
        if hit: sl[v][0] += 1
    w.close()
print(f"[2/6] slot: {len(sl)} 家，覆盖 {len(slot_dates)} 天（{min(slot_dates) if slot_dates else '-'} ~ {max(slot_dates) if slot_dates else '-'}）")

# ---------- 3. credit：当月各周快照 逾期>7天占比 均值（不含坏账） ----------
def load_debt(path):
    out = {}
    if path.endswith(".xlsx"):
        w = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = w.worksheets[0]; ws.reset_dimensions()
        for r in ws.iter_rows(values_only=True):
            if not r or r[0] is None: continue
            s0 = str(r[0])
            if s0.startswith("#") or "VENDOR" in s0.upper() or "CÓDIGO" in s0.upper() or is_agg(s0): continue
            out[norm(r[0])] = dict(m7=abs(num(r[3]) or 0), o7=abs(pct(r[4]) or 0))
        w.close()
    else:
        rr = list(csv.reader(io.open(path, encoding="utf-8-sig")))
        hd = [x.strip() for x in rr[0]]
        i7 = hd.index("Mora >7 días") if "Mora >7 días" in hd else 3
        io7 = hd.index("% mora >7d") if "% mora >7d" in hd else 4
        for r in rr[1:]:
            if not r or not r[0].strip() or is_agg(r[0]): continue
            out[norm(r[0])] = dict(m7=abs(num(r[i7]) or 0), o7=abs(pct(r[io7]) or 0))
    return out
snaps = []
for f in MONTH["debts"]:
    p = R + f
    if os.path.exists(p): snaps.append(load_debt(p))
    else: issues.append({"type": "debt_file_missing", "file": f})
print(f"[3/6] credit: {len(snaps)} 期周快照（只用逾期>7天占比，坏账已剔除）")

# ---------- 4+6. 日表：D-3R% 与 完单骑手（自然月） ----------
dd = defaultdict(lambda: dict(assigned=0.0, r3=0.0, complete=0.0, perfect=0.0, couriers=set()))
perfect_col_seen = set()
bill_dates = set()
for f in MONTH["bills"]:
    p = R + r"周度账单&成本\周度账单" + "\\" + f
    if not os.path.exists(p): issues.append({"type": "bill_missing", "file": f}); continue
    w = openpyxl.load_workbook(p, read_only=True, data_only=True)
    if "courier by day" not in w.sheetnames: w.close(); continue
    ws = w["courier by day"]; ws.reset_dimensions(); rows = ws.iter_rows(values_only=True)
    hh = [str(x).strip() if x else "" for x in next(rows)]
    jx = {n: i for i, n in enumerate(hh) if n}
    # 完美单列名带口径后缀（7 月为 "Perfect_Orders (10min)"，8 月起切回 1min），按前缀定位
    _pc = [n for n in jx if n.startswith("Perfect_Orders")]
    if len(_pc) != 1:
        issues.append({"type": "perfect_col_ambiguous", "file": f, "found": _pc}); w.close(); continue
    i_perf = jx[_pc[0]]; perfect_col_seen.add(_pc[0])
    for r in rows:
        if not r or r[jx["Vendor_Code"]] is None: continue
        d = str(r[jx["Date"]])[:10]
        if not in_month(d): continue
        bill_dates.add(d)
        a = dd[norm(r[jx["Vendor_Code"]])]
        a["assigned"] += num(r[jx["Assigned_Orders"]]) or 0
        a["complete"] += num(r[jx["Complete_Orders"]]) or 0
        a["perfect"]  += num(r[i_perf]) or 0
        a["r3"] += sum(num(r[jx[k]]) or 0 for k in ("D_Duty_Reject", "D_Duty_Reschedule", "D_Duty_Reassign"))
        cid = r[jx["Courier_ID"]]
        if cid is not None and (num(r[jx["Complete_Orders"]]) or 0) > 0:
            a["couriers"].add(str(cid).strip())
    w.close()
print(f"[4/6] d3r + 完单 + 完美单: {len(dd)} 家，覆盖 {len(bill_dates)} 天｜完美单列 {sorted(perfect_col_seen)}")
if len(perfect_col_seen) > 1:
    issues.append({"type": "perfect_caliber_mixed", "cols": sorted(perfect_col_seen),
                   "note": "当期内完美单口径发生切换，月度汇总跨口径，需人工确认"})

# ---------- 5. newrider：brand_new_curp ÷ new_riders ----------
w = openpyxl.load_workbook(R + MONTH["newrider"], read_only=True, data_only=True)
ws = w.worksheets[0]; ws.reset_dimensions(); rows = ws.iter_rows(values_only=True)
hh = [str(x).strip() if x else "" for x in next(rows)]
jx = {n: i for i, n in enumerate(hh) if n}
NEW = {}
for r in rows:
    if not r or r[0] is None or is_agg(r[0]): continue
    NEW[norm(r[0])] = dict(new_riders=num(r[jx["new_riders"]]) or 0,
                           brand_new=num(r[jx["brand_new_curp"]]) or 0,
                           period=str(r[jx["period"]]) if "period" in jx else "")
w.close()
print(f"[5/6] newrider: {len(NEW)} 家（分母=new_riders）")

# ---------- 6. blocked：AF 红线封禁去重骑手 ----------
w = openpyxl.load_workbook(R + MONTH["blocks"], read_only=True, data_only=True)
ws = w.worksheets[0]; ws.reset_dimensions(); rows = ws.iter_rows(values_only=True)
hh = [str(x).replace("\ufeff", "").strip() if x else "" for x in next(rows)]
jx = {n: i for i, n in enumerate(hh) if n}
BAN = defaultdict(set); nrec = 0
for r in rows:
    if not r or r[jx["vendor_code"]] is None or is_agg(r[jx["vendor_code"]]): continue
    if str(r[jx["block_strategy_name"]]).strip() != "反作弊-红线封禁": continue
    d = str(r[jx["real_block_begin_time_local"]])[:10]
    if not in_month(d): continue
    nrec += 1
    BAN[norm(r[jx["vendor_code"]])].add(str(r[jx["rider_id"]]).strip())
w.close()
print(f"[6/6] blocked: {nrec} 条 → {len(BAN)} 家有封禁")

# ---------- 组装 ----------
OUT = []
for v, d in V.items():
    iss = []
    # 日均完美单量 = 账单 courier by day 的 Perfect_Orders 自然月合计 ÷ 当月自然日数
    _a = dd.get(v)
    pdaily = (_a["perfect"] / DAYS) if _a else None
    if pdaily is None: iss.append("orders_missing")
    s = sl.get(v)
    slot_rate = (s[0] / s[1] * 100) if s and s[1] else None
    if slot_rate is None: iss.append("slot_missing")
    o7s = [snap[v]["o7"] for snap in snaps if v in snap and snap[v]["m7"] > 0]
    no_debt = len(o7s) == 0
    overdue_ratio = (sum(o7s) / len(o7s)) if o7s else 0.0
    a = dd.get(v)
    d3r = (a["r3"] / a["assigned"] * 100) if a and a["assigned"] else None
    if d3r is None: iss.append("d3r_missing")
    cc = len(a["couriers"]) if a else 0
    if cc == 0: iss.append("no_completing_couriers")
    n = NEW.get(v)
    newr = (n["brand_new"] / n["new_riders"] * 100) if n and n["new_riders"] else None
    if n is None: iss.append("newrider_missing")
    elif not n["new_riders"]: iss.append("no_new_riders")   # 当月没招人 → 无法计算比率
    banned = len(BAN.get(v, ()))
    blk = (banned / cc * 100) if cc else None
    OUT.append({
        "vendor_code": d["code"], "vendor_key": v, "city": d["city"],
        "official_level_v1": d["official_level"],
        "values": {
            K_ORD: (round(pdaily, 2) if pdaily is not None else None),
            K_SLOT: (round(slot_rate, 2) if slot_rate is not None else None),
            K_CRED: {"overdue_ratio": round(overdue_ratio, 2), "no_debt": no_debt},
            K_D3R: (round(d3r, 3) if d3r is not None else None),
            K_NEW: (round(newr, 2) if newr is not None else None),
            K_BLK: (round(blk, 2) if blk is not None else None),
        },
        "raw": {
            "perfect_orders": (round(a["perfect"]) if a else None), "days_in_period": DAYS,
            "completed_orders": (round(a["complete"]) if a else None),
            "result_sheet_daily_completed": d["daily"],           # 交叉校验用，非计分口径
            "result_sheet_perfect_col": (round(d["perfect_month"]) if d["perfect_month"] is not None else None),
            "slot_achieved": (s[0] if s else None), "slot_target": (s[1] if s else None),
            "credit_snapshots": len(o7s), "assigned_orders": (round(a["assigned"]) if a else None),
            "d3r_events": (round(a["r3"]) if a else None), "completing_couriers": cc,
            "new_riders": (n["new_riders"] if n else None), "brand_new_curp": (n["brand_new"] if n else None),
            "blocked_riders_dedup": banned,
        },
        "issues": iss,
    })

meta = {
    "period": MONTH["period"], "period_label": MONTH["label"], "period_type": "monthly",
    "days_in_period": DAYS, "vendor_count": len(OUT),
    "change_request": "CR-20260814",
    "caliber": {
        "orders": f"周账单 courier by day.Perfect_Orders 自然月合计 ÷ {DAYS}（真完美单口径；⚠️ 7 月为 10 分钟口径，8 月起切 1 分钟）。商分级结果.xlsx 的 Perfect order 列名不实（=完单量）且仅 37/102 家有值，不可用，见 BUG-003",
        "slot": "slot_detail 日表按自然月过滤：Σ达成 slot ÷ Σ目标 slot（单 slot 达成阈值 70%，已用周表校验 113/113 吻合）",
        "credit": "当月各周逾期明细快照的「逾期>7天占比」均值；坏账占比不再参与（CR-20260814 B1）",
        "d3r": "courier by day 按自然月过滤：Σ(Reject+Reschedule+Reassign) ÷ Σ Assigned_Orders（公式误差 0.000pp 已验证）",
        "newrider": "brand_new_curp ÷ new_riders（分母改为当月新招骑手，CR-20260814 B3）",
        "blocked_rider_rate": "当月 AF 红线封禁去重骑手 ÷ 当月完单骑手（日表 Courier_ID 去重）",
    },
    "slot_dates_covered": sorted(slot_dates), "bill_dates_covered": sorted(bill_dates),
    "data_issues": issues,
}
p = os.path.join(ROOT, "data", f"_monthly_values_{MONTH['period']}.json")
io.open(p, "w", encoding="utf-8").write(json.dumps({"meta": meta, "vendors": OUT}, ensure_ascii=False, indent=1))
print(f"\n取数完成 → {p}")
for k, lab in [("slot_missing", "Slot 缺"), ("d3r_missing", "3R 缺"), ("newrider_missing", "拉新缺"),
               ("no_new_riders", "当月未招新人"), ("no_completing_couriers", "无完单骑手")]:
    c = sum(1 for o in OUT if k in o["issues"])
    if c: print(f"  ⚠️ {lab}: {c} 家")
if issues: print("  数据源问题:", issues)
