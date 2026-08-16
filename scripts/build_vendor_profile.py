# -*- coding: utf-8 -*-
"""
build_vendor_profile.py — 生成 data/vendor_profile.json

输入：
  01_原始数据/商基础信息/首单时间@0815.xlsx     → 首单日期（2026-08-16 起改用此版；
                                                 源头已按拼写变体合并，多一列 code_variants）
  01_原始数据/商基础信息/Vendor_List_2026-08-05.xlsx → 显示名 / 城市 / 经营状态
  05_网站/data/rm_assignment.json              → RM 归属（已处理，不重新解析 xlsx）
  05_网站/data/vendors_2026_07.json            → 102 家主名单（vendor_code 以此为准）

规则（见 CLAUDE.md 铁律 5/6/15）：
  - 主键 = 完整 vendor_code 全串，禁止数字后缀匹配
  - 匹配前归一化：去重音 + 转大写 + 去首尾空白
  - 对不上的列入 unmatched 报告，不猜、不补
  - 邮箱/电话/联系人/银行账号等个人信息一律不写入输出
"""
import json
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent.parent  # 商分级/
SITE = ROOT / "05_网站"
RAW = ROOT / "01_原始数据" / "商基础信息"


def normalize(code):
    """去重音 + 转大写 + 去空白。已知案例：ENVÍAGUIA/ENVIAGUIA、ALEJANDRO/Alejandro、MONDRAGÓN..."""
    if code is None:
        return None
    s = unicodedata.normalize("NFKD", str(code).strip())
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return s.upper()


def load_first_orders():
    """同一 vendor 在源表里按拼写变体拆成多行（vendor_id 相同、日期范围不同）。
    首单日期取同组最早值。若同一归一化键下 vendor_id 不一致则不合并，记入 conflicts。"""
    wb = openpyxl.load_workbook(RAW / "首单时间@0815.xlsx", read_only=True, data_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    header = list(rows[0])
    idx = {h: i for i, h in enumerate(header)}
    groups = {}
    for r in rows[1:]:
        if r is None or r[idx["vendor_code"]] is None:
            continue
        groups.setdefault(normalize(r[idx["vendor_code"]]), []).append(r)

    out = {}
    multi_row = []
    conflicts = []
    for key, g in groups.items():
        ids = {r[idx["vendor_id"]] for r in g}
        if len(ids) > 1:
            conflicts.append({
                "normalized_code": key,
                "rows": [{"raw": str(r[idx["vendor_code"]]).strip(),
                          "vendor_id": str(r[idx["vendor_id"]]),
                          "first_order_date": r[idx["first_order_date"]].date().isoformat()} for r in g],
            })
            continue
        dates = sorted(r[idx["first_order_date"]] for r in g if isinstance(r[idx["first_order_date"]], datetime))
        if not dates:
            continue
        earliest = dates[0].date().isoformat()
        canonical_raw = max((str(r[idx["vendor_code"]]).strip() for r in g), key=lambda s: sum(c.isupper() for c in s))
        if len(g) > 1:
            multi_row.append({
                "normalized_code": key,
                "row_count": len(g),
                "vendor_id": str(next(iter(ids))),
                "first_order_dates": [d.date().isoformat() for d in dates],
                "used": earliest,
            })
        out[key] = {
            "vendor_code_raw": canonical_raw,
            "city": g[0][idx["city_name"]],
            "first_order_date": earliest,
        }
    return out, multi_row, conflicts


def load_vendor_list():
    wb = openpyxl.load_workbook(RAW / "Vendor_List_2026-08-05.xlsx", read_only=True, data_only=True)
    ws = wb["Vendor清单"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    header = list(rows[0])
    idx = {h: i for i, h in enumerate(header)}
    out = {}
    for r in rows[1:]:
        code = r[idx["Vendor Code"]]
        if code is None:
            continue
        key = normalize(code)
        out[key] = {
            "vendor_code_raw": str(code).strip(),
            "display_name": (r[idx["供应商名称"]] or "").strip() or None,
            "city": r[idx["城市"]],
            "active_status": r[idx["是否活跃"]],
        }
    return out


def main():
    vendors = json.loads((SITE / "data" / "vendors_2026_07.json").read_text(encoding="utf-8"))
    rm_data = json.loads((SITE / "data" / "rm_assignment.json").read_text(encoding="utf-8"))
    rm_by_code = {normalize(a["vendor_code"]): a["rm"] for a in rm_data["assignments"]}
    rm_unassigned = {normalize(c) for c in rm_data.get("unmatched_vendors", [])}

    first_orders, fo_multi_row, fo_conflicts = load_first_orders()
    vendor_list = load_vendor_list()

    profiles = []
    missing_in_list = []
    missing_first_order = []
    city_conflicts = []
    nonexact_matches = []

    for v in vendors["vendors"]:
        code_raw = v["vendor_code"]
        key = normalize(code_raw)
        vl = vendor_list.get(key)
        fo = first_orders.get(key)

        if vl is None:
            missing_in_list.append(code_raw)
        elif vl["vendor_code_raw"] != code_raw:
            nonexact_matches.append({"canonical": code_raw, "vendor_list_raw": vl["vendor_code_raw"]})
        if fo is None:
            missing_first_order.append(code_raw)
        elif fo["vendor_code_raw"] != code_raw:
            nonexact_matches.append({"canonical": code_raw, "first_order_raw": fo["vendor_code_raw"]})

        for src_name, src in (("vendor_list", vl), ("first_order", fo)):
            if src and src.get("city") and src["city"] != v["city"]:
                city_conflicts.append({
                    "vendor_code": code_raw, "计算表": v["city"],
                    "source": src_name, "source_city": src["city"],
                })

        rm = rm_by_code.get(key)
        profiles.append({
            "vendor_code": code_raw,
            "display_name": vl["display_name"] if vl else None,
            "city": v["city"],
            "rm": rm,
            "rm_status": "assigned" if rm else ("unassigned" if key in rm_unassigned or vl or fo else "unknown"),
            "first_order_date": fo["first_order_date"] if fo else None,
            "active_status": vl["active_status"] if vl else None,
        })

    out = {
        "_readme": "由 scripts/build_vendor_profile.py 生成。与 vendors_2026_07.json 通过完整 vendor_code 关联。不含邮箱/电话/联系人/银行账号等个人信息。",
        "_generated_from": {
            "display_name": "Vendor_List_2026-08-05.xlsx · 供应商名称",
            "city": "vendors_2026_07.json · city（与源表冲突时以测算表为准，冲突列于 _issues.city_conflicts）",
            "rm": "data/rm_assignment.json（源：RM映射关系@0731.xlsx）",
            "first_order_date": "Vendor首单时间.xlsx · first_order_date",
            "active_status": "Vendor_List_2026-08-05.xlsx · 是否活跃",
        },
        "_generated_at": "2026-08-13",
        "_match_rule": "完整 vendor_code，归一化（NFKD 去重音 + 大写 + 去空白）后比对；禁止数字后缀匹配",
        "vendor_count": len(profiles),
        "profiles": profiles,
        "_issues": {
            "missing_in_vendor_list": missing_in_list,
            "missing_first_order": missing_first_order,
            "city_conflicts": city_conflicts,
            "nonexact_code_spellings": nonexact_matches,
            "rm_unassigned": sorted([p["vendor_code"] for p in profiles if p["rm"] is None]),
            "first_order_multi_row_groups": {
                "_note": "首单时间源表中同一 vendor（vendor_id 相同）按拼写变体拆成多行，首单日期取最早值。",
                "groups": fo_multi_row,
            },
            "first_order_vendor_id_conflicts": fo_conflicts,
        },
    }

    dest = SITE / "data" / "vendor_profile.json"
    dest.write_text(json.dumps(out, ensure_ascii=False, indent=1), encoding="utf-8")

    print(f"profiles: {len(profiles)}")
    print(f"缺显示名(不在Vendor_List): {len(missing_in_list)} -> {missing_in_list}")
    print(f"缺首单日期: {len(missing_first_order)} -> {missing_first_order}")
    print(f"城市冲突: {len(city_conflicts)} -> {city_conflicts}")
    print(f"写法不一致(归一化后才匹配上): {len(nonexact_matches)} -> {nonexact_matches}")
    print(f"无RM: {out['_issues']['rm_unassigned']}")
    print(f"首单表多行聚合(取最早): {len(fo_multi_row)} 组")
    print(f"首单表 vendor_id 冲突(未合并): {len(fo_conflicts)} -> {fo_conflicts}")
    named = sum(1 for p in profiles if p["display_name"])
    dated = sum(1 for p in profiles if p["first_order_date"])
    print(f"有显示名: {named}/102, 有首单日期: {dated}/102")


if __name__ == "__main__":
    sys.exit(main())
