# -*- coding: utf-8 -*-
"""
周度取数层（只取数 + 清洗，不算分）
=====================================
算分一律交给 src/engine（唯一实现）。本脚本产出引擎可直接消费的原始值。

输出：data/_weekly_values_<period>.json
  { meta: {...}, vendors: [ { vendor_code, city, values: {...}, raw: {...}, issues: [...] } ] }

口径定义见 docs/指标取数口径.md（项目根 docs/），本脚本不做任何档位/权重/分数线判断。

用法：改 WEEK 配置后 `python scripts/extract_weekly.py`
"""
import openpyxl, csv, io, json, unicodedata, os
from collections import Counter

WEEK = dict(
    period="2026-W32",
    label="wk0803（2026-08-03 ~ 08-09）",
    days=7,
    src=r"C:\Users\abbychang\Desktop\商分级\01_原始数据\0803-0809",
    bill="0810Bill-wk0803.xlsx",
    debt="逾期明细_0811.csv",
    newr="0803-0809新骑手.xlsx",
    comp="账号合规.xlsx",
)
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# 指标 key 从 rules.json 动态读取——rules.json 是唯一事实来源，任一侧改名两边都不会崩。
# （2026-08-13 实证：r2→d3r、identity→blocked_rider_rate 改名后硬编码 key 会全量 value_missing）
_rules = json.load(io.open(os.path.join(ROOT, "data", "rules.json"), encoding="utf-8"))
_KEYS = [i["key"] for i in _rules["indicators"]]
def _key(*candidates):
    for c in candidates:
        if c in _KEYS: return c
    raise SystemExit(f"rules.json 中找不到指标 key，候选={candidates}，现有={_KEYS}")
K_ORDERS = _key("orders")
K_SLOT   = _key("slot")
K_CREDIT = _key("credit")
K_D3R    = _key("d3r", "r2")                      # D-3R%（账单列名 3R，7月表列名 D-duty 2R%）
K_NEW    = _key("newrider")
K_BLOCK  = _key("blocked_rider_rate", "identity")  # 合规账号率（值=封禁率，lower_better）

def norm(s):
    """vendor_code 归一化：去重音 + 转大写。主键用完整全串，禁用数字后缀。"""
    s = str(s).strip().upper()
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
def num(x):
    try: return float(str(x).replace("%", "").replace(",", ""))
    except: return None
def pct(x):
    v = num(x)
    return None if v is None else (v * 100 if abs(v) <= 1.5 else v)
def is_agg(code):
    c = str(code).strip().upper()
    return c.startswith("__") or c in ("TOTAL", "合计") or "TOTAL" in c

D = WEEK["src"] + "\\"
# Slot 日表目录：01_原始数据/周度账单&成本/Slot达成率（由 WEEK["src"] 的父目录推导，避免写死绝对路径）
SLOT_DIR = os.path.join(os.path.dirname(WEEK["src"]), "周度账单&成本", "Slot达成率")
issues_global = []

# ---------- 账单：完美单 / 3R / 完单骑手 / 城市 / Slot ----------
wb = openpyxl.load_workbook(D + WEEK["bill"], read_only=True, data_only=True)
ws = wb["Vendor_Bill"]; ws.reset_dimensions()
V = {}; hdr = None
for r in ws.iter_rows(values_only=True):
    if hdr is None:
        if r and r[0] == "Vendor_Code":
            hdr = list(r); ix = {n: i for i, n in enumerate(hdr) if n}
        continue
    if not r or r[0] is None or is_agg(r[0]): continue
    v = norm(r[0])
    city = str(r[ix["City"]]).strip().upper() if "City" in ix and r[ix["City"]] else ""
    if city not in ("CDMX", "MTY"):
        city = "MTY" if "MTY" in v else "CDMX"
    V[v] = dict(code=str(r[0]).strip(), city=city,
                perfect=num(r[ix["Perfect_Order"]]) or 0,
                delivered=num(r[ix["Delivered_Order"]]) or 0,
                r3=pct(r[ix["3R"]]),
                cc=num(r[ix["Complete_Courier"]]) or 0,
                slot=None)
nm = next((x for x in wb.sheetnames if x.lower().startswith("slot")), None)
slot_only = []
if nm:
    ws = wb[nm]; ws.reset_dimensions(); h2 = None
    for r in ws.iter_rows(values_only=True):
        if h2 is None:
            if r and any(x and str(x).strip() == "Vendor" for x in r):
                h2 = [str(x).strip() if x else "" for x in r]
                vi = h2.index("Vendor"); si = next(i for i, x in enumerate(h2) if "w/o surge" in x)
            continue
        if not r or r[vi] is None or is_agg(r[vi]): continue
        v = norm(r[vi]); val = num(r[si])
        if val is None: continue
        val = val * 100 if val <= 1.5 else val
        if v in V: V[v]["slot"] = val
        else: slot_only.append(str(r[vi]).strip())
wb.close()

# CR-20260814 B2：Slot 改为按日聚合 Σ达成 ÷ Σ总 slot 数
# 判定列用 "slot ach?"（0/1 真实达成）；"slot ach final" 含周末/比赛加权，只影响全职奖金额，不进等级考核
if WEEK.get("slot_detail"):
    sp = os.path.join(SLOT_DIR, WEEK["slot_detail"])
    w2 = openpyxl.load_workbook(sp, read_only=True, data_only=True)
    nm2 = next((x for x in w2.sheetnames if "detail" in x.lower() and "rider" not in x.lower()), None)
    if nm2:
        ws2 = w2[nm2]; ws2.reset_dimensions(); rr = ws2.iter_rows(values_only=True)
        hh = [str(x).strip() if x else "" for x in next(rr)]
        jx = {n: i for i, n in enumerate(hh) if n}
        ivd = jx.get("Vendor Code") or jx.get("Vendor code")
        itg, iok, iach = jx.get("Target"), jx.get("slot ach?"), jx.get("achieve_rider_cnt")
        agg = {}
        for r in rr:
            if not r or ivd is None or r[ivd] is None or is_agg(r[ivd]): continue
            tgt = num(r[itg]) if itg is not None else None
            if tgt is None or tgt <= 0: continue
            v = norm(r[ivd]); a = agg.setdefault(v, [0, 0]); a[1] += 1
            hit = ((num(r[iok]) or 0) >= 1) if (iok is not None and r[iok] is not None) \
                  else (((num(r[iach]) or 0) / tgt) >= 0.70)
            if hit: a[0] += 1
        for v, a in agg.items():
            if v in V and a[1]:
                V[v]["slot"] = a[0] / a[1] * 100
                V[v]["slot_achieved"], V[v]["slot_total"] = a[0], a[1]
    w2.close()
if slot_only:
    issues_global.append({"type": "slot_only_no_bill", "detail": "Slot 表有但 Vendor_Bill 无", "vendors": slot_only})

# ---------- 欠款 ----------
rows = list(csv.reader(io.open(D + WEEK["debt"], encoding="utf-8-sig")))
dh = [x.strip() for x in rows[0]]
i_m7, i_o7, i_o30 = dh.index("Mora >7 días"), dh.index("% mora >7d"), dh.index("% mora >30d")
DEBT = {}
for r in rows[1:]:
    if not r or not r[0].strip() or is_agg(r[0]): continue
    DEBT[norm(r[0])] = dict(code=r[0].strip(), m7=abs(num(r[i_m7]) or 0),
                            o7=abs(pct(r[i_o7]) or 0), o30=abs(pct(r[i_o30]) or 0))
debt_only = [d["code"] for v, d in DEBT.items() if v not in V]
if debt_only:
    issues_global.append({"type": "debt_only_no_bill", "detail": "欠款表有但本周账单无（多为已退出仍欠款商）", "count": len(debt_only), "vendors": debt_only})

# ---------- 新骑手（CURP 行业全新） ----------
wb = openpyxl.load_workbook(D + WEEK["newr"], read_only=True, data_only=True)
ws = wb.worksheets[0]; ws.reset_dimensions(); rows = ws.iter_rows(values_only=True)
nh = [str(x).strip() if x else "" for x in next(rows)]
i_bn = nh.index("brand_new_curp")
i_nr = nh.index("new_riders")     # CR-20260814 B3：分母改为当期新招骑手
NEW = {}; new_only = []
for r in rows:
    if not r or r[0] is None or is_agg(r[0]): continue
    v = norm(r[0])
    NEW[v] = dict(brand_new=num(r[i_bn]) or 0, new_riders=num(r[i_nr]) or 0)
    if v not in V: new_only.append(str(r[0]).strip())
wb.close()
if new_only:
    issues_global.append({"type": "newrider_only_no_bill", "detail": "新骑手表有但账单无", "vendors": new_only})

# ---------- 合规账号（AF 红线封禁，rider_id 去重） ----------
wb = openpyxl.load_workbook(D + WEEK["comp"], read_only=True, data_only=True)
ws = wb.worksheets[0]; ws.reset_dimensions(); rows = ws.iter_rows(values_only=True)
ch = [str(x).strip() if x else "" for x in next(rows)]
i_v, i_rid = ch.index("vendor_code"), ch.index("rider_id")
i_cs, i_bs = ch.index("channel_source"), ch.index("block_strategy_name")
BAN = {}; rec_total = 0; ban_only = []
for r in rows:
    if not r or r[i_v] is None or is_agg(r[i_v]): continue
    if str(r[i_cs]).strip() != "Anti-Fraud": continue
    if str(r[i_bs]).strip() != "反作弊-红线封禁": continue
    rec_total += 1
    v = norm(r[i_v])
    BAN.setdefault(v, set()).add(str(r[i_rid]).strip())
    if v not in V and str(r[i_v]).strip() not in ban_only: ban_only.append(str(r[i_v]).strip())
wb.close()
if ban_only:
    issues_global.append({"type": "compliance_only_no_bill", "detail": "封禁表有但账单无", "vendors": ban_only})

# ---------- 组装（只出原始值，不算分） ----------
OUT = []
for v, d in V.items():
    issues = []
    pdaily = d["perfect"] / WEEK["days"]
    if d["slot"] is None: issues.append("slot_missing")
    if d["r3"] is None: issues.append("r3_missing")
    if v not in NEW: issues.append("newrider_missing")
    if d["cc"] == 0: issues.append("complete_courier_zero")   # 拉新/合规无法计算
    dbt = DEBT.get(v)
    banned = len(BAN.get(v, ()))
    _nr = NEW.get(v)
    newr_rate = (_nr["brand_new"] / _nr["new_riders"] * 100) if (_nr and _nr["new_riders"]) else None
    ban_rate = (banned / d["cc"] * 100) if d["cc"] else None
    OUT.append({
        "vendor_code": d["code"],
        "vendor_key": v,
        "city": d["city"],
        # 引擎消费的原始值（key 全部取自 rules.json，见文件头 _key()）
        "values": {
            K_ORDERS: round(pdaily, 2),          # 日均完美单量（当期完美单 ÷ 天数）
            K_SLOT: d["slot"],                   # Slot 达成率 %（w/o surge）
            K_CREDIT: {                          # 复合指标：两分项原始值，引擎内部 scoreComposite
                "overdue_ratio": (dbt["o7"] if dbt and dbt["m7"] > 0 else 0.0),
                "bad_debt_ratio": (dbt["o30"] if dbt and dbt["m7"] > 0 else 0.0),
                "no_debt": (dbt is None or dbt["m7"] == 0),
            },
            K_D3R: d["r3"],                      # D-3R%（账单列名 3R）
            K_NEW: (round(newr_rate, 2) if newr_rate is not None else None),
            K_BLOCK: (round(ban_rate, 2) if ban_rate is not None else None),
        },
        # 原始计数，供页面展示与追溯
        "raw": {
            "perfect_orders": round(d["perfect"]),
            "delivered_orders": round(d["delivered"]),
            "complete_couriers": round(d["cc"]),
            "brand_new_curp": (_nr["brand_new"] if _nr else None),
            "new_riders": (_nr["new_riders"] if _nr else None),
            "blocked_riders_dedup": banned,
            "overdue_7d_amount": (dbt["m7"] if dbt else 0.0),
        },
        # 红线：周度只能判「单周命中」，触发需整月
        "redline_week_hit": bool(dbt and dbt["m7"] >= 50000 and dbt["o7"] >= 50),
        "issues": issues,
    })

meta = {
    "period": WEEK["period"],
    "period_label": WEEK["label"],
    "period_type": "weekly",
    "days_in_period": WEEK["days"],
    "vendor_count": len(OUT),
    "disclaimer": "周度试算，非正式评级，不计入保护期与清退累计。等级按自然月评定。",
    "redline_note": "周度仅判定 redline_week_hit（单周命中）；redline_triggered 需当月命中≥2 次且月末最后一周仍命中，单周数据无法判定。",
    "flex_adjustment": "本期不含灵活分。",
    "sources": {
        "orders/r2/complete_couriers/slot": WEEK["bill"],
        "credit": WEEK["debt"],
        "newrider": WEEK["newr"] + "（分母用账单 Complete_Courier）",
        "identity": WEEK["comp"] + "（channel_source=Anti-Fraud 且 block_strategy_name=反作弊-红线封禁，rider_id 去重）",
    },
    "compliance_records_filtered": rec_total,
    "data_issues": issues_global,
}
out_path = os.path.join(ROOT, "data", f"_weekly_values_{WEEK['period']}.json")
io.open(out_path, "w", encoding="utf-8").write(json.dumps({"meta": meta, "vendors": OUT}, ensure_ascii=False, indent=1))
print(f"取数完成：{len(OUT)} 家 → {out_path}")
print(f"  Slot 缺失 {sum(1 for o in OUT if 'slot_missing' in o['issues'])} | 新骑手缺失 {sum(1 for o in OUT if 'newrider_missing' in o['issues'])} | 完单骑手为0 {sum(1 for o in OUT if 'complete_courier_zero' in o['issues'])}")
print(f"  单周红线命中 {sum(1 for o in OUT if o['redline_week_hit'])} 家 | 封禁记录 {rec_total} 条")
for it in issues_global:
    print(f"  [对不上] {it['type']}: {it.get('count', len(it.get('vendors', [])))} 家 — {it['detail']}")
