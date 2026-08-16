# -*- coding: utf-8 -*-
"""English vendor-level report for RM leadership (from _report_en_2026-07.json)"""
import json, io, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
d = json.load(io.open(os.path.join(ROOT, "data", "_report_en_2026-07.json"), encoding="utf-8"))
rm = json.load(io.open(os.path.join(ROOT, "data", "rm_assignment.json"), encoding="utf-8"))
import unicodedata
def norm(s):
    s = str(s).strip().upper()
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
RM = {norm(a["vendor_code"]): a.get("rm") for a in rm.get("assignments", [])}

wb = openpyxl.Workbook()
bold = Font(bold=True); white = Font(bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="333333")
S_fill = PatternFill("solid", fgColor="FFF3D6"); down_fill = PatternFill("solid", fgColor="FDE8E8")
up_fill = PatternFill("solid", fgColor="E9F4E9"); warn = PatternFill("solid", fgColor="FDF0EC")

# ---------- Sheet 1: Summary ----------
ws = wb.active; ws.title = "Summary"
j, w = d["july"], d["week"]
rows = [
    ["Vendor Level Simulation — New Rules (CR-20260814)"], [],
    ["Prepared", d["meta"]["generated"], "", "Engine", d["meta"]["engine"]], [],
    ["⚠️ This is a SIMULATION on actual data, not an official rating. Do not share with vendors."], [],
    ["JULY 2026 (Jul 1–31, natural month) — the reference period"], [],
    ["City", "S", "A", "B", "C", "Total", "S+A share", "Thresholds"],
    ["CDMX", j["by_city"]["CDMX"]["S"], j["by_city"]["CDMX"]["A"], j["by_city"]["CDMX"]["B"], j["by_city"]["CDMX"]["C"],
     j["by_city"]["CDMX"]["n"], f'{j["by_city"]["CDMX"]["sa"]:.0f}%', "S≥80 / A≥55 / C<15"],
    ["MTY", j["by_city"]["MTY"]["S"], j["by_city"]["MTY"]["A"], j["by_city"]["MTY"]["B"], j["by_city"]["MTY"]["C"],
     j["by_city"]["MTY"]["n"], f'{j["by_city"]["MTY"]["sa"]:.0f}%', "S≥90 / A≥65 / C<15"],
    ["Total", j["total"]["S"], j["total"]["A"], j["total"]["B"], j["total"]["C"], j["total"]["n"], f'{j["total"]["sa"]:.0f}%', ""],
    [],
    ["Monthly incentive (MXN)", j["monthly_incentive_mxn"], "", "vs current payout", ""],
    [],
    ["Scoring model", "Daily perfect orders 30% + Slot achievement 25% + Repayment credit 20% + D-3R% 15% + New rider ratio 5% + Account compliance 5%"],
    ["New rider bands", "≥70% = 100 / 50–70% = 80 / 30–50% = 50 / ≤30% = 0"],
    ["Zero-score bands", "All zero-score bands are inclusive of the boundary: daily perfect orders ≤100, slot ≤40%, "
     "overdue share ≥50%, D-3R% ≥3%, new rider ratio ≤30%, blocked rider rate ≥15%"],
    ["Repayment credit bands", "no overdue debt = 100 / ≤20% = 80 / 20–50% = 50 / ≥50% = 0 (share of debt overdue >7 days)"],
    [],
    ["WEEK 2026-W32 (Aug 3–9) — trend check only"], [],
    ["City", "S", "A", "B", "C", "Total", "S+A share"],
    ["CDMX", w["by_city"]["CDMX"]["S"], w["by_city"]["CDMX"]["A"], w["by_city"]["CDMX"]["B"], w["by_city"]["CDMX"]["C"], w["by_city"]["CDMX"]["n"], f'{w["by_city"]["CDMX"]["sa"]:.0f}%'],
    ["MTY", w["by_city"]["MTY"]["S"], w["by_city"]["MTY"]["A"], w["by_city"]["MTY"]["B"], w["by_city"]["MTY"]["C"], w["by_city"]["MTY"]["n"], f'{w["by_city"]["MTY"]["sa"]:.0f}%'],
    ["Total", w["total"]["S"], w["total"]["A"], w["total"]["B"], w["total"]["C"], w["total"]["n"], f'{w["total"]["sa"]:.0f}%'],
    [],
    ["🚫 WHY THE WEEKLY NUMBERS LOOK MUCH BETTER — DO NOT COMPARE THEM WITH THE MONTHLY RESULT"],
    ["", "Weekly (W32)", "Monthly (July)", "Why"],
    ["Vendors with no overdue debt", "68%", "21%", "Weekly uses ONE snapshot; monthly averages 4 weekly snapshots — any overdue week counts"],
    ["Vendors with zero blocked riders", "69%", "21%", "Far fewer anti-fraud events fit into a single week"],
    ["Avg. daily perfect orders", "612", "407", "Weekly covers only the 93 vendors that billed that week; monthly includes 102 incl. dormant ones"],
    [],
    ["→ Use the JULY figures for any decision. The weekly view is for spotting trends and red-line hits only."],
]
for r in rows: ws.append(r)
# 行号按内容定位，避免增删行时错位（旧版硬编码行号已错位一行）
def row_at(prefix):
    for i, r in enumerate(rows, 1):
        if r and isinstance(r[0], str) and r[0].startswith(prefix): return i
    raise KeyError(prefix)
ws["A1"].font = Font(bold=True, size=14)
ws.cell(row=row_at("⚠️"), column=1).font = Font(bold=True, color="C00000")
for p in ("JULY 2026", "WEEK 2026-W32"):
    ws.cell(row=row_at(p), column=1).font = Font(bold=True, size=12)
for p in ("City", "Total"):
    for i, r in enumerate(rows, 1):
        if r and r[0] == p:
            for cell in ws[i]: cell.font = bold
ws.cell(row=row_at("🚫"), column=1).font = Font(bold=True, color="C00000")
for cell in ws[row_at("🚫") + 1]: cell.font = bold
ws.cell(row=row_at("→ Use the JULY"), column=1).font = Font(bold=True, color="1F6F1F")
ws.column_dimensions["A"].width = 34
for c in "BCDEFG": ws.column_dimensions[c].width = 14
ws.column_dimensions["H"].width = 22
ws.column_dimensions["D"].width = 20

# ---------- Sheet 2: Vendor detail ----------
ws2 = wb.create_sheet("Vendor Detail (July)")
H = ["Vendor code", "City", "RM", "Level (new rules)", "Level (current rules)", "Change", "Score",
     "Weakest metric", "Daily perfect orders", "Slot %", "Overdue >7d %", "D-3R %", "New rider %",
     "Blocked rider %", "Perfect orders (month)", "New riders", "Brand-new riders", "Blocked riders",
     "Level W32", "Score W32"]
ws2.append(H)
for cell in ws2[1]:
    cell.font = white; cell.fill = hdr_fill; cell.alignment = Alignment(wrap_text=True, vertical="center")
for v in d["vendors"]:
    m = v["metrics"]; raw = v["raw"]
    ws2.append([v["vendor_code"], v["city"], RM.get(norm(v["vendor_code"])) or "Unassigned",
                v["level_july_new_rules"], v["level_july_current_rules"] or "—", v["change_vs_current"],
                v["score_july"], v["weakest_metric"], m["daily_perfect_orders"], m["slot_pct"],
                ("no debt" if m["no_debt"] else m["overdue_ratio_pct"]), m["d3r_pct"],
                (m["new_rider_pct"] if m["new_rider_pct"] is not None else "no new hires"),
                m["blocked_rider_pct"], raw.get("perfect_orders"), raw.get("new_riders"),
                raw.get("brand_new_curp"), raw.get("blocked_riders_dedup"),
                v["level_week_2026W32"] or "—", v["score_week"] if v["score_week"] is not None else "—"])
    r = ws2.max_row
    if v["level_july_new_rules"] == "S":
        for cell in ws2[r]: cell.fill = S_fill
    elif v["change_vs_current"] == "DOWN":
        ws2.cell(row=r, column=6).fill = down_fill
    elif v["change_vs_current"] == "UP":
        ws2.cell(row=r, column=6).fill = up_fill
ws2.freeze_panes = "D2"
ws2.auto_filter.ref = f"A1:{get_column_letter(len(H))}{ws2.max_row}"
for i, wd in enumerate([26, 7, 18, 15, 16, 9, 8, 26, 13, 9, 12, 9, 11, 12, 15, 11, 14, 13, 11, 11]):
    ws2.column_dimensions[get_column_letter(i + 1)].width = wd

# ---------- Sheet 3: By RM ----------
ws3 = wb.create_sheet("By RM")
ws3.append(["RM", "Vendors", "S", "A", "B", "C", "S+A share", "Vendors moving DOWN vs current rules"])
for cell in ws3[1]: cell.font = white; cell.fill = hdr_fill
agg = {}
for v in d["vendors"]:
    k = RM.get(norm(v["vendor_code"])) or "Unassigned"
    a = agg.setdefault(k, {"n": 0, "S": 0, "A": 0, "B": 0, "C": 0, "down": []})
    a["n"] += 1; a[v["level_july_new_rules"]] += 1
    if v["change_vs_current"] == "DOWN": a["down"].append(v["vendor_code"])
for k, a in sorted(agg.items(), key=lambda kv: -kv[1]["n"]):
    ws3.append([k, a["n"], a["S"], a["A"], a["B"], a["C"], f'{(a["S"]+a["A"])/a["n"]*100:.0f}%', len(a["down"])])
for i, wd in enumerate([20, 9, 6, 6, 6, 6, 11, 34]): ws3.column_dimensions[get_column_letter(i + 1)].width = wd
ws3.freeze_panes = "A2"

# ---------- Sheet 4: Methodology ----------
ws4 = wb.create_sheet("Methodology")
for r in [["Methodology & data sources"], [],
 ["Metric", "Definition", "Source"],
 ["Daily perfect orders (30%)", "Perfect_Orders from the weekly bill (courier by day), summed over the calendar month ÷ calendar days (31 for July). ⚠️ July is on the 10-minute definition; the metric switched to 1 minute from August, so the band thresholds must be re-checked before August is scored.", "周度账单 courier by day.Perfect_Orders"],
 ["Slot achievement (25%)", "Slots achieved ÷ total target slots, accumulated daily over the month. A slot counts as achieved when valid attending riders ≥70% of target (verified: matches the weekly report for 113/113 vendors). Weekend/match weighting affects full-time bonuses only, not levels.", "Slot达成率/slot_detail*.xlsx (daily)"],
 ["Repayment credit (20%)", "Share of debt overdue >7 days, taken as a weekly snapshot and averaged over the month. Bands: no overdue debt = 100 / ≤20% = 80 / 20–50% = 50 / ≥50% = 0. Full marks are reserved for vendors with no overdue debt at all. Bad debt (>30 days) no longer contributes to the score.", "欠款情况/逾期明细*.csv"],
 ["D-3R% (15%)", "(Reject + Reschedule + Reassign) ÷ Assigned orders, daily rows filtered to the calendar month (formula verified to 0.000pp against the weekly bill)", "Bill / courier by day"],
 ["New rider ratio (5%)", "Riders recruited in the month who are brand new to the industry (CURP-verified) ÷ total riders recruited that month", "骑手拉新/7月新骑手.xlsx"],
 ["Account compliance (5%)", "Distinct riders blocked by anti-fraud red-line ÷ riders with completed orders that month (lower is better)", "人证合一/5-8月封禁明细.xlsx"],
 [], ["Levels", "S / A / C thresholds are set per city. Levels depend on the composite score only — no volume gate.", ""],
 ["Red line", "Weekly hit = debt overdue >7d ≥ MXN 50,000 AND ≥50% of total debt. Triggered when hit in ≥2 weeks of the month AND still hit in the final week → cash incentive forfeited, level capped at B. The weekly view can only show single-week hits, never a trigger.", ""],
 ["Not included", "Flexible adjustment (±10 points) is not applied in this simulation.", ""],
 [], ["Known data gaps", "8 vendors recruited no riders in July → new rider ratio cannot be computed, scored 0. 21 vendors appear in the Slot file but not in the weekly bill. 34 vendors appear in the debt file but had no bill that week (mostly exited vendors still in arrears).", ""]]:
    ws4.append(r)
ws4["A1"].font = Font(bold=True, size=13)
for cell in ws4[3]: cell.font = bold
ws4.column_dimensions["A"].width = 28; ws4.column_dimensions["B"].width = 105; ws4.column_dimensions["C"].width = 32
for row in ws4.iter_rows(min_row=4):
    for cell in row: cell.alignment = Alignment(wrap_text=True, vertical="top")

p = os.path.join(os.path.dirname(ROOT), "03_汇报材料", "Vendor Level Simulation - 2026-08-14 (EN).xlsx")
wb.save(p)
print("saved:", p)
print(f"vendors {len(d['vendors'])} | RM matched {sum(1 for v in d['vendors'] if RM.get(norm(v['vendor_code'])))}")
