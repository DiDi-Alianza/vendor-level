# -*- coding: utf-8 -*-
"""
多周原始值提取 —— 用于「周度是否天然虚高」的对照分析。

产出 data/_weekly_values_2026-W{nn}.json（只有原始值，不含任何档位判断/算分）。
算分由 scripts/score_weeks.js 调 src/engine/rules.js 完成 —— 不在本文件重复实现。

口径与 extract_monthly.py 完全一致，只把「自然月」换成「周区间（周一~周日）」：
  orders   = courier by day.Perfect_Orders 区间合计 ÷ 7
  slot     = Slot 日表区间内 Σ达成 ÷ Σ目标（判定列 "slot ach?"，0/1 真实达成）
  d3r      = (Reject+Reschedule+Reassign) ÷ Assigned_Orders
  credit   = 该周欠款快照的「逾期>7天占比」（周度只有单点，无月内均值）
  blocked  = 反作弊红线封禁骑手(去重，区间内) ÷ 区间完单骑手
  newrider = ⚠️ 固定取 7 月整月值。周度招人基数太小（几个人就能让比例剧烈跳动），
             逐周取会引入噪声。固定住它 = 把这一项从周间差异中剔除，
             使对比只反映其余五项的真实变化。权重仅 5%。

用法：PYTHONIOENCODING=utf-8 python scripts/extract_weeks.py
"""
import openpyxl, csv, io, json, os, unicodedata, datetime
from collections import defaultdict

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = r"C:\Users\abbychang\Desktop\商分级\01_原始数据"
BILLS = os.path.join(SRC, "周度账单&成本", "周度账单")
SLOTS = os.path.join(SRC, "周度账单&成本", "Slot达成率")
BLOCKS = os.path.join(SRC, "人证合一", "5-8月封禁明细.xlsx")

# (周号, 周一, 周日, 账单, Slot 文件, 欠款快照)
WEEKS = [
    ("2026-W27", "2026-06-29", "2026-07-05", "0706Bill-wk0629_adj.xlsx",   "Slot_detail_0629.xlsx",
     r"欠款情况\0629-0705.xlsx"),
    ("2026-W28", "2026-07-06", "2026-07-12", "0713Bill-wk0706-adj.xlsx",   "Slot_Result wk0706.xlsx",
     r"欠款情况\0706-0712\逾期指标_0715.csv"),
    ("2026-W29", "2026-07-13", "2026-07-19", "0720Bill-wk0713_adj.xlsx",   "Slot wk0713.xlsx",
     r"欠款情况\0713-0719\逾期指标_0721.csv"),
    ("2026-W30", "2026-07-20", "2026-07-26", "0727Bill-wk0720_adj.xlsx",   "slot_detail_0720.xlsx",
     r"欠款情况\0720-0726\逾期明细_0729.csv"),
    ("2026-W31", "2026-07-27", "2026-08-02", "0803Bill-wk0727 - adj.xlsx", "slot_detail0727.xlsx",
     r"欠款情况\0727-0802\逾期明细_0803.csv"),
    ("2026-W32", "2026-08-03", "2026-08-09", "0810Bill-wk0803.xlsx",       "slot_detail0803-Lea.xlsx",
     r"欠款情况\0803-0809\逾期明细_0811.csv"),
]
# W32 的账单不在周账单目录下
BILL_OVERRIDE = {"0810Bill-wk0803.xlsx": os.path.join(SRC, "0803-0809", "0810Bill-wk0803.xlsx")}


def norm(s):
    s = str(s).strip().upper()
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def num(x):
    if x is None: return None
    try: return float(str(x).replace("%", "").replace(",", "").strip())
    except ValueError: return None


def pct(x):
    v = num(x)
    return None if v is None else (v * 100 if abs(v) <= 1.5 else v)


def is_agg(c):
    u = str(c).strip().upper()
    return u.startswith("#") or "TOTAL" in u or "合计" in u or u in ("VENDOR_CODE", "VENDOR", "CÓDIGO")


def in_range(d, a, b):
    return a <= d[:10] <= b


def load_bill(path, a, b, issues):
    agg = defaultdict(lambda: dict(perfect=0.0, assigned=0.0, r3=0.0, couriers=set()))
    days, cols = set(), set()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "courier by day" not in wb.sheetnames:
        issues.append({"type": "no_courier_sheet", "file": os.path.basename(path)}); wb.close(); return agg, days, cols
    ws = wb["courier by day"]; ws.reset_dimensions(); it = ws.iter_rows(values_only=True)
    h = [str(x).strip() if x else "" for x in next(it)]
    jx = {n: i for i, n in enumerate(h) if n}
    pc = [n for n in jx if n.startswith("Perfect_Orders")]
    if len(pc) != 1:
        issues.append({"type": "perfect_col_ambiguous", "found": pc}); wb.close(); return agg, days, cols
    ip = jx[pc[0]]; cols.add(pc[0])
    for r in it:
        if not r or r[jx["Vendor_Code"]] is None: continue
        d = str(r[jx["Date"]])[:10]
        if not in_range(d, a, b): continue
        days.add(d)
        x = agg[norm(r[jx["Vendor_Code"]])]
        x["perfect"] += num(r[ip]) or 0
        x["assigned"] += num(r[jx["Assigned_Orders"]]) or 0
        x["r3"] += sum(num(r[jx[k]]) or 0 for k in ("D_Duty_Reject", "D_Duty_Reschedule", "D_Duty_Reassign"))
        cid = r[jx["Courier_ID"]]
        if cid is not None and (num(r[jx["Complete_Orders"]]) or 0) > 0:
            x["couriers"].add(str(cid).strip())
    wb.close()
    return agg, days, cols


def load_slot(path, a, b, issues):
    sl = defaultdict(lambda: [0, 0])
    dates = set()
    if not os.path.exists(path):
        issues.append({"type": "slot_file_missing", "file": os.path.basename(path)}); return sl, dates
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    nm = next((x for x in wb.sheetnames if "detail" in x.lower() and "rider" not in x.lower()), None)
    if not nm:
        issues.append({"type": "slot_sheet_missing", "file": os.path.basename(path),
                       "sheets": wb.sheetnames}); wb.close(); return sl, dates
    ws = wb[nm]; ws.reset_dimensions(); it = ws.iter_rows(values_only=True)
    h = [str(x).strip() if x else "" for x in next(it)]
    jx = {n: i for i, n in enumerate(h) if n}
    ivd = jx.get("Vendor Code") or jx.get("Vendor code")
    idt, itg, iach, ifin = jx.get("Date"), jx.get("Target"), jx.get("achieve_rider_cnt"), jx.get("slot ach?")
    if ivd is None or idt is None or itg is None:
        issues.append({"type": "slot_cols_missing", "file": os.path.basename(path), "header": h[:20]})
        wb.close(); return sl, dates
    for r in it:
        if not r or r[ivd] is None or is_agg(r[ivd]): continue
        d = str(r[idt])[:10]
        if not in_range(d, a, b): continue
        dates.add(d)
        tgt = num(r[itg])
        if tgt is None or tgt <= 0: continue
        v = norm(r[ivd]); sl[v][1] += 1
        if ifin is not None and r[ifin] is not None:
            ok = num(r[ifin])
            hit = (ok == 1) if ok is not None else (str(r[ifin]).strip().upper() in ("Y", "YES", "TRUE"))
        else:
            hit = ((num(r[iach]) or 0) / tgt) >= 0.70
        if hit: sl[v][0] += 1
    wb.close()
    return sl, dates


def load_debt(path, issues):
    out = {}
    if not os.path.exists(path):
        issues.append({"type": "debt_file_missing", "file": os.path.basename(path)}); return out
    if path.endswith(".xlsx"):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.worksheets[0]; ws.reset_dimensions()
        for r in ws.iter_rows(values_only=True):
            if not r or r[0] is None or is_agg(r[0]): continue
            out[norm(r[0])] = dict(m7=abs(num(r[3]) or 0), o7=abs(pct(r[4]) or 0))
        wb.close()
    else:
        rr = list(csv.reader(io.open(path, encoding="utf-8-sig")))
        hd = [x.strip() for x in rr[0]]
        i7 = hd.index("Mora >7 días") if "Mora >7 días" in hd else 3
        io7 = hd.index("% mora >7d") if "% mora >7d" in hd else 4
        for r in rr[1:]:
            if not r or not r[0].strip() or is_agg(r[0]): continue
            out[norm(r[0])] = dict(m7=abs(num(r[i7]) or 0), o7=abs(pct(r[io7]) or 0))
    return out


def load_blocks(a, b, issues):
    out = defaultdict(set)
    if not os.path.exists(BLOCKS):
        issues.append({"type": "blocks_file_missing"}); return out
    wb = openpyxl.load_workbook(BLOCKS, read_only=True, data_only=True)
    ws = wb.worksheets[0]; ws.reset_dimensions(); it = ws.iter_rows(values_only=True)
    h = [str(x).strip() if x else "" for x in next(it)]
    jx = {n: i for i, n in enumerate(h) if n}

    def col(*c):
        for x in c:                      # 先精确匹配
            for k in jx:
                if k.lower() == x.lower(): return jx[k]
        for x in c:                      # 再子串匹配
            for k in jx:
                if x.lower() in k.lower(): return jx[k]
        return None

    iv, ic = col("vendor_code", "vendor"), col("rider_id", "courier_id", "骑手id")
    ich, ibs, idt = col("channel_source"), col("block_strategy_name"), col("dt", "date", "时间")
    if None in (iv, ic):
        issues.append({"type": "blocks_cols_missing", "header": h}); wb.close(); return out
    for r in it:
        if not r or r[iv] is None: continue
        if ich is not None and "ANTI-FRAUD" not in str(r[ich]).upper(): continue
        if ibs is not None and "红线" not in str(r[ibs]): continue
        if idt is not None and not in_range(str(r[idt]), a, b): continue
        out[norm(r[iv])].add(str(r[ic]).strip())
    wb.close()
    return out


# 拉新固定取 7 月整月值（见文件头说明）
JUL = {v["vendor_key"]: v["values"]["newrider"]
       for v in json.load(io.open(os.path.join(ROOT, "data", "_monthly_values_2026-07.json"),
                                  encoding="utf-8"))["vendors"]}

for wk, a, b, bill, slotf, debtf in WEEKS:
    issues = []
    bp = BILL_OVERRIDE.get(bill, os.path.join(BILLS, bill))
    if not os.path.exists(bp):
        print(f"[{wk}] ⛔ 账单不存在：{bill}"); continue
    agg, days, cols = load_bill(bp, a, b, issues)
    sl, sdates = load_slot(os.path.join(SLOTS, slotf), a, b, issues)
    debt = load_debt(os.path.join(SRC, debtf), issues)
    ban = load_blocks(a, b, issues)
    ndays = (datetime.date.fromisoformat(b) - datetime.date.fromisoformat(a)).days + 1
    if len(days) != ndays:
        issues.append({"type": "days_incomplete", "expected": ndays, "got": len(days)})

    vendors = []
    for v, x in sorted(agg.items()):
        cc = len(x["couriers"])
        s = sl.get(v)
        dt = debt.get(v)
        vendors.append({
            "vendor_code": v, "vendor_key": v,
            "city": "MTY" if "MTY" in v else "CDMX",
            "values": {
                "orders": round(x["perfect"] / ndays, 2),
                "slot": (round(s[0] / s[1] * 100, 2) if s and s[1] else None),
                "credit": {"overdue_ratio": (round(dt["o7"], 2) if dt and dt["m7"] > 0 else 0.0),
                           "no_debt": not (dt and dt["m7"] > 0)},
                "d3r": (round(x["r3"] / x["assigned"] * 100, 3) if x["assigned"] else None),
                "newrider": JUL.get(v),
                "blocked_rider_rate": (round(len(ban.get(v, ())) / cc * 100, 2) if cc else None),
            },
            "raw": {
                "perfect_orders": round(x["perfect"]), "days_in_period": ndays,
                "assigned_orders": round(x["assigned"]), "d3r_events": round(x["r3"]),
                "completing_couriers": cc,
                "slot_achieved": (s[0] if s else None), "slot_target": (s[1] if s else None),
                "blocked_riders_dedup": len(ban.get(v, ())),
                "debt_overdue_7d_mxn": (round(dt["m7"]) if dt else None),
            },
        })
    out = {
        "meta": {
            "period": wk, "period_type": "weekly", "period_label": f"{wk}（{a} ~ {b}）",
            "days_in_period": ndays, "vendor_count": len(vendors),
            "week_start": a, "week_end": b,
            "sources": {"bill": bill, "slot": slotf, "debt": debtf, "blocks": os.path.basename(BLOCKS)},
            "caliber_note": "newrider 固定取 2026-07 整月值（周度基数过小），其余五项均为本周实测。"
                            "credit 为单周快照，无月内均值。红线只能判「单周命中」，不能判「触发」。",
            "perfect_col": sorted(cols),
            "days_covered": len(days), "slot_days_covered": len(sdates),
        },
        "issues": issues,
        "vendors": vendors,
    }
    p = os.path.join(ROOT, "data", f"_weekly_values_{wk}.json")
    io.open(p, "w", encoding="utf-8").write(json.dumps(out, ensure_ascii=False, indent=1))
    slot_n = sum(1 for v in vendors if v["values"]["slot"] is not None)
    print(f"[{wk}] {a}~{b}  {len(vendors):>3} 家｜账单 {len(days)}/{ndays} 天｜Slot {slot_n} 家/{len(sdates)} 天"
          f"｜欠款 {len(debt)} 家｜封禁 {len(ban)} 家｜{cols}｜issues {len(issues)}")
    for i in issues[:4]:
        print("    ⚠️", json.dumps(i, ensure_ascii=False)[:130])
